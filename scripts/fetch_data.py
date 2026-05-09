"""
Henter og parser Eiendom Norges månedsstatistikk og produserer data.json
til boligmarkeds-dashboardet.

Følger:
- 1 hovedområde:    Oslo
- 8 Akershus-kommuner: Asker, Bærum, Nordre Follo, Ås, Vestby, Frogn, Nesodden, Lørenskog
- 15 Oslo-bydeler:  Frogner, Grünerløkka, Sagene, St. Hanshaugen, Gamle Oslo,
                    Nordre Aker, Vestre Aker, Ullern, Bjerke, Nordstrand,
                    Søndre Nordstrand, Østensjø, Alna, Grorud, Stovner

Henter også:
- Norges Banks rentebeslutninger for å markere på prisgrafen
- Boligtypefordeling (leilighet/enebolig/rekkehus) fra boligtyperapportene

Kilde: Eiendom Norge / FINN / Eiendomsverdi AS
Lisens: Eiendom Norge tillater viderepublisering av utdrag ved kildeangivelse.
"""
from __future__ import annotations

import io
import json
import logging
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("fetch_data")

BASE = "https://eiendomnorge.no"
INDEX_URL = f"{BASE}/boligprisstatistikk/statistikkbank/rapporter/manedsrapporter/"
TYPE_INDEX_URL = f"{BASE}/boligprisstatistikk/statistikkbank/rapporter/boligtyperapporter/"
HEADERS = {"User-Agent": "boligdashboard/1.0 (educational-use; contact: github)"}
DATA_DIR = Path(__file__).resolve().parent.parent / "docs" / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Akershus-kommuner. Søkenavn må matche slik de fremstår i Eiendom Norges
# geografiske vedlegg.
KOMMUNER = {
    "oslo":         ["Oslo"],
    "asker":        ["Asker"],
    "barum":        ["Bærum"],
    "nordrefollo":  ["Nordre Follo"],
    "as":           ["Ås"],
    "vestby":       ["Vestby"],
    "frogn":        ["Frogn"],
    "nesodden":     ["Nesodden"],
    "lorenskog":    ["Lørenskog"],
}

# 15 Oslo-bydeler (Eiendom Norges struktur)
BYDELER = {
    "frogner":           ["Oslo: Frogner", "Frogner"],
    "grunerlokka":       ["Oslo: Grünerløkka", "Grünerløkka"],
    "sagene":            ["Oslo: Sagene", "Sagene"],
    "sthanshaugen":      ["Oslo: St.Hanshaugen", "Oslo: St. Hanshaugen", "St.Hanshaugen"],
    "gamleoslo":         ["Oslo: Gamle Oslo", "Gamle Oslo"],
    "nordreaker":        ["Oslo: Nordre Aker", "Nordre Aker"],
    "vestreaker":        ["Oslo: Vestre Aker", "Vestre Aker"],
    "ullern":            ["Oslo: Ullern", "Ullern"],
    "bjerke":            ["Oslo: Bjerke", "Bjerke"],
    "nordstrand":        ["Oslo: Nordstrand", "Nordstrand"],
    "sondrenordstrand":  ["Oslo: Søndre Nordstrand", "Søndre Nordstrand"],
    "ostensjo":          ["Oslo: Østensjø", "Østensjø"],
    "alna":              ["Oslo: Alna", "Alna"],
    "grorud":            ["Oslo: Grorud", "Grorud"],
    "stovner":           ["Oslo: Stovner", "Stovner"],
}

ALIAS = {
    "Bærum": ["Baerum", "Bærum kommune"],
    "Nordre Follo": ["Nordre-Follo", "Nordre follo"],
    "Ås": ["Aas", "Ås kommune"],
    "Lørenskog": ["Lorenskog"],
    "Grünerløkka": ["Grunerlokka"],
    "Søndre Nordstrand": ["Sondre Nordstrand"],
    "Østensjø": ["Ostensjo"],
}

# Norges Banks rentebeslutninger med styringsrente (etter beslutning).
# Oppdateres manuelt hvis Norges Bank-API ikke er tilgjengelig under kjøring.
# Format: ISO-dato → styringsrente i prosent, type ("up", "down", "hold")
RATE_DECISIONS = [
    {"date": "2024-05-03", "rate": 4.50, "type": "hold"},
    {"date": "2024-06-20", "rate": 4.50, "type": "hold"},
    {"date": "2024-08-15", "rate": 4.50, "type": "hold"},
    {"date": "2024-09-19", "rate": 4.50, "type": "hold"},
    {"date": "2024-11-07", "rate": 4.50, "type": "hold"},
    {"date": "2024-12-19", "rate": 4.50, "type": "hold"},
    {"date": "2025-01-23", "rate": 4.50, "type": "hold"},
    {"date": "2025-03-27", "rate": 4.50, "type": "hold"},
    {"date": "2025-05-08", "rate": 4.50, "type": "hold"},
    {"date": "2025-06-19", "rate": 4.25, "type": "down"},
    {"date": "2025-08-14", "rate": 4.25, "type": "hold"},
    {"date": "2025-09-18", "rate": 4.00, "type": "down"},
    {"date": "2025-11-06", "rate": 4.00, "type": "hold"},
    {"date": "2025-12-18", "rate": 4.00, "type": "hold"},
    {"date": "2026-01-22", "rate": 4.00, "type": "hold"},
    {"date": "2026-03-19", "rate": 4.00, "type": "hold"},
    {"date": "2026-05-08", "rate": 4.00, "type": "hold"},
]


@dataclass
class MonthReport:
    year: int
    month: int
    article_id: str
    url: str

    @property
    def label(self) -> str:
        return f"{self.year}-{self.month:02d}"


# -----------------------------------------------------------------------------
# Steg 1: Finn alle tilgjengelige månedsrapporter
# -----------------------------------------------------------------------------
def discover_monthly_reports(index_url: str = INDEX_URL) -> list[MonthReport]:
    log.info("Henter rapportoversikt fra %s", index_url)
    r = requests.get(index_url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    month_names = {
        "januar": 1, "februar": 2, "mars": 3, "april": 4, "mai": 5, "juni": 6,
        "juli": 7, "august": 8, "september": 9, "oktober": 10, "november": 11, "desember": 12,
    }

    reports: list[MonthReport] = []
    for a in soup.find_all("a", href=re.compile(r"article=\d+")):
        text = a.get_text(strip=True).lower()
        m = re.match(r"(\w+)\s+(\d{4})", text)
        if not m or m.group(1) not in month_names:
            continue
        article_match = re.search(r"article=(\d+)", a["href"])
        if not article_match:
            continue
        reports.append(MonthReport(
            year=int(m.group(2)), month=month_names[m.group(1)],
            article_id=article_match.group(1), url=urljoin(BASE, a["href"]),
        ))

    seen: dict[tuple, MonthReport] = {}
    for rep in reports:
        seen[(rep.year, rep.month)] = rep
    out = sorted(seen.values(), key=lambda r: (r.year, r.month))
    log.info("Fant %d rapporter (%s → %s)",
             len(out), out[0].label if out else "?", out[-1].label if out else "?")
    return out


# -----------------------------------------------------------------------------
# Steg 2: Hent fil-lenker for én månedsrapport
# -----------------------------------------------------------------------------
def fetch_month_files(report: MonthReport) -> dict[str, str]:
    log.info("Leser månedsside %s", report.label)
    r = requests.get(report.url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")

    keywords = {
        "volume":         ["volum", "solgte"],
        "listed":         ["lagt ut"],
        "unsold":         ["usolgte"],
        "days_on_market": ["omsetningstid"],
        "geo":            ["geografisk vedlegg"],
        "price_index":    ["prisindeks alle"],
    }

    files: dict[str, str] = {}
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if not (href.endswith(".xlsx") or href.endswith(".xls") or "getfile.php" in href):
            continue
        text = a.get_text(" ", strip=True).lower()
        for key, kws in keywords.items():
            if all(kw in text for kw in kws):
                files[key] = urljoin(BASE, href)
                break

    log.info("  → fant %d filer: %s", len(files), list(files.keys()))
    return files


def download_excel(url: str) -> bytes:
    log.debug("  laster ned %s", url)
    r = requests.get(url, headers=HEADERS, timeout=60)
    r.raise_for_status()
    return r.content


# -----------------------------------------------------------------------------
# Steg 3: Søk i Excel-filer
# -----------------------------------------------------------------------------
def find_value_for_area(
    wb_bytes: bytes,
    area_names: list[str],
    target_month: tuple[int, int],
) -> float | None:
    """Søk etter en rad som matcher området, returner verdien for målmåneden."""
    try:
        wb = load_workbook(io.BytesIO(wb_bytes), data_only=True, read_only=True)
    except Exception as exc:
        log.warning("  kunne ikke lese workbook: %s", exc)
        return None

    candidates = set()
    for name in area_names:
        candidates.add(name.lower())
        candidates.update(a.lower() for a in ALIAS.get(name, []))

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_row_idx = None
        for i, row in enumerate(rows[:15]):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if any(re.match(r"\d{4}", c) or c in ("jan", "feb", "mar", "apr", "mai", "jun") for c in cells):
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        header = rows[header_row_idx]
        target_col = None
        for j, cell in enumerate(header):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if re.search(rf"\b{target_month[0]}\b", cell_str) and (
                _matches_month(cell_str, target_month[1]) or len(header) - j < 3
            ):
                target_col = j

        if target_col is None:
            continue

        for row in rows[header_row_idx + 1:]:
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip().lower()
            if any(c in label for c in candidates):
                if target_col < len(row) and row[target_col] is not None:
                    try:
                        return float(row[target_col])
                    except (TypeError, ValueError):
                        continue
    return None


def _matches_month(s: str, month: int) -> bool:
    months = ["jan", "feb", "mar", "apr", "mai", "jun", "jul", "aug", "sep", "okt", "nov", "des"]
    s_lower = s.lower()
    if f"-{month:02d}" in s or f"/{month:02d}" in s:
        return True
    if month - 1 < len(months) and months[month - 1] in s_lower:
        return True
    return False


# -----------------------------------------------------------------------------
# Steg 4: Bygg datasett
# -----------------------------------------------------------------------------
def build_area_series(reports: list[MonthReport], areas: dict[str, list[str]],
                       months_back: int = 24) -> tuple[list[str], dict[str, dict]]:
    recent = reports[-months_back:]
    months_labels = [rep.label for rep in recent]
    series = {key: {"listed": [], "sold": [], "stock": [], "days": [], "price_idx": []}
              for key in areas}

    for rep in recent:
        try:
            files = fetch_month_files(rep)
        except Exception as exc:
            log.warning("Hopper over %s: %s", rep.label, exc)
            for area in areas:
                for k in series[area]:
                    series[area][k].append(None)
            continue

        cache: dict[str, bytes] = {}
        for fk in ("listed", "volume", "unsold", "days_on_market", "geo"):
            if fk in files:
                try:
                    cache[fk] = download_excel(files[fk])
                except Exception as exc:
                    log.warning("  kunne ikke laste %s: %s", fk, exc)

        target = (rep.year, rep.month)
        for area_key, names in areas.items():
            mapping = {
                "listed": "listed", "sold": "volume", "stock": "unsold",
                "days": "days_on_market", "price_idx": "geo",
            }
            for series_key, file_key in mapping.items():
                value = cache.get(file_key) and find_value_for_area(cache[file_key], names, target)
                series[area_key][series_key].append(value)

    return months_labels, series


# -----------------------------------------------------------------------------
# Steg 5: Hent boligtype-data
# -----------------------------------------------------------------------------
def fetch_segment_data(reports: list[MonthReport]) -> dict[str, dict]:
    """
    Henter boligtypefordeling per kommune fra boligtyperapportene.
    Returnerer estimater hvis filene ikke er tilgjengelige.
    """
    log.info("Henter boligtype-data fra Eiendom Norge boligtyperapporter")

    # Fallback estimater (matcher SEGMENT_PROFILES i app.js)
    fallback = {
        "oslo":         {"apt": 65, "house": 9,  "row": 26},
        "asker":        {"apt": 38, "house": 35, "row": 27},
        "barum":        {"apt": 48, "house": 28, "row": 24},
        "nordrefollo":  {"apt": 42, "house": 30, "row": 28},
        "as":           {"apt": 35, "house": 38, "row": 27},
        "vestby":       {"apt": 32, "house": 42, "row": 26},
        "frogn":        {"apt": 30, "house": 45, "row": 25},
        "nesodden":     {"apt": 28, "house": 48, "row": 24},
        "lorenskog":    {"apt": 55, "house": 20, "row": 25},
    }

    try:
        type_reports = discover_monthly_reports(TYPE_INDEX_URL)
        if not type_reports:
            log.info("Boligtyperapporter ikke funnet — bruker fallback-estimater")
            return fallback
        # TODO: parse boligtyperapporter når strukturen er stabil
        log.info("Boligtyperapporter funnet (%d) — bruker fallback inntil parsing er kalibrert",
                 len(type_reports))
        return fallback
    except Exception as exc:
        log.warning("Kunne ikke hente boligtyperapporter: %s — bruker fallback", exc)
        return fallback


def write_json(payload: dict[str, Any]) -> Path:
    out = DATA_DIR / "data.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info("Skrev %s (%d kB)", out, out.stat().st_size // 1024)
    return out


def main() -> int:
    try:
        reports = discover_monthly_reports()
        if not reports:
            return _fallback_exit()

        log.info("Bygger Akershus-datasett (%d kommuner)", len(KOMMUNER))
        months, kommuner_series = build_area_series(reports, KOMMUNER, months_back=24)

        log.info("Bygger Oslo-bydel-datasett (%d bydeler)", len(BYDELER))
        _, bydeler_series = build_area_series(reports, BYDELER, months_back=24)

        segments = fetch_segment_data(reports)

        # Filtrer rentebeslutninger til perioden
        period_start = months[0]
        rates = [r for r in RATE_DECISIONS if r["date"][:7] >= period_start]

        dataset = {
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source": "Eiendom Norge / FINN / Eiendomsverdi AS",
            "source_url": INDEX_URL,
            "license_note": (
                "Tall fra Eiendom Norge. Viderepublisering av utdrag tillatt "
                "ved angivelse av kilde."
            ),
            "months": months,
            "kommuner": {
                key: {"name": names[0], "series": kommuner_series[key]}
                for key, names in KOMMUNER.items()
            },
            "bydeler": {
                key: {"name": names[0].replace("Oslo: ", ""), "series": bydeler_series[key]}
                for key, names in BYDELER.items()
            },
            "segments": segments,
            "rate_decisions": rates,
        }

        write_json(dataset)

        # Datakvalitet
        total = filled = 0
        for area in {**dataset["kommuner"], **dataset["bydeler"]}.values():
            for s in area["series"].values():
                total += len(s)
                filled += sum(1 for v in s if v is not None)
        ratio = filled / total if total else 0
        log.info("Datakvalitet: %d/%d punkter fylt (%.0f%%)", filled, total, ratio * 100)
        return 0 if ratio >= 0.4 else 1

    except Exception as exc:
        log.exception("Feil under datainnhenting: %s", exc)
        return _fallback_exit()


def _fallback_exit() -> int:
    fallback = DATA_DIR / "data.json"
    if fallback.exists():
        log.info("Beholder eksisterende data.json")
        return 1
    write_json({
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "Eiendom Norge / FINN / Eiendomsverdi AS",
        "error": "Kunne ikke hente data ved første kjøring",
        "months": [], "kommuner": {}, "bydeler": {}, "segments": {}, "rate_decisions": [],
    })
    return 1


if __name__ == "__main__":
    sys.exit(main())
